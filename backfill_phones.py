"""backfill_phones.py — Import phone numbers from Numbers column in XLSX."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl, sqlite3, re
from pathlib import Path

XLSX = Path(__file__).parent / "kochi_software_companies.xlsx"
DB   = Path.home() / '.jobtracker' / 'data.db'

def clean_phone(val):
    if not val:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'unknown', 'n/a', '-', ''):
        return None
    # Keep digits, +, spaces, hyphens, parens
    cleaned = re.sub(r'[^\d\s+\-()/]', '', s).strip()
    return cleaned if cleaned else None

wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
ws = wb.active

headers = [str(h).replace('\ufeff','').strip() if h else '' for h in next(ws.iter_rows(values_only=True))]
print("Headers:", headers)

name_idx   = headers.index('Company Name')
number_idx = headers.index('Numbers') if 'Numbers' in headers else None
print(f"Name col: {name_idx}, Numbers col: {number_idx}")

con = sqlite3.connect(str(DB))
cur = con.cursor()

updated = 0
skipped = 0
for row in ws.iter_rows(values_only=True, min_row=2):
    name  = str(row[name_idx]).strip() if row[name_idx] else None
    phone = clean_phone(row[number_idx]) if number_idx is not None else None

    if not name or not phone:
        skipped += 1
        continue

    cur.execute(
        "UPDATE companies SET phone_number = ? WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (phone, name)
    )
    if cur.rowcount:
        updated += 1
    else:
        skipped += 1

con.commit()
con.close()
wb.close()

print(f"\nPhone numbers imported: {updated}")
print(f"Skipped (no phone / no match): {skipped}")

# Show sample
con = sqlite3.connect(str(DB))
cur = con.cursor()
cur.execute("SELECT name, phone_number FROM companies WHERE phone_number IS NOT NULL LIMIT 10")
print("\nSample companies with phones:")
for row in cur.fetchall():
    print(f"  {row[0][:40]:40s}  {row[1]}")
con.close()
